import os
from openpyxl import load_workbook
from openpyxl import Workbook
import re
os.chdir(r"d:/ll")
wb2 = load_workbook("test1.xlsx")
#print wb2.sheetnames
for i in range (1,wb2["all"].max_row+1):
    car_Number = wb2["all"][str("H"+str(i))].value
    print car_Number
    print type(car_Number)
    car_Filter = re.search(r'(?<=浙)\w',car_Number)
    print car_Filter.group(0)
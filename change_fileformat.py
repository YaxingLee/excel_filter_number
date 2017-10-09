#coding = utf-8
import xlrd
import os
import re
from openpyxl.workbook import Workbook
import winsound
os.chdir(r"D:\excel\3k6k\30")
def cvt_xls_to_xlsx(src_file_path, dst_file_path):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()
    sheet_names = book_xls.sheet_names()
    for sheet_index in range(0,len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)
    book_xlsx.save(dst_file_path)
    print "converted %s" %src_file_path
    os.remove(src_file_path)
    print "deleted %s" %src_file_path

name_List = os.listdir(".")
src_Name = []
dst_Name = []
for i in name_List:
    try:
        name = re.search(r'xls$', i).group(0)
        if name:
            src_Name.append(i)
            dst_Name.append(i+"x")
        else:
            pass
    except:
        pass
#print src_Name
if len(src_Name)==len(dst_Name):
    print "ok"
erroList = []
for a in range(0,len(src_Name)):
    if dst_Name[a] in os.listdir("."):
        print src_Name[a], "already have xlsx format."
        os.remove(src_Name[a])
        print "deleted ",src_Name[a]
#        cvt_xls_to_xlsx(src_Name[a],dst_Name[a])
    else:
#        print "converting ",src_Name[a]
        try:
            cvt_xls_to_xlsx(src_Name[a], dst_Name[a])
        except:
            erroList.append(src_Name[a])
print "all done!!!"
print erroList
winsound.Beep(600,1000)
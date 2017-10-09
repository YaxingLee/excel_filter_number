import os
from openpyxl import load_workbook
from openpyxl import Workbook
import re
import copy
from tel_filter import tel_filter
os.chdir(r"d:\ll")
def get_sheet(file):
    xls_File = file
    wb2 = load_workbook(xls_File)
    print wb2.sheetnames
def filter_Number(file,sheet_Name):
    Tel_Ningbo = tel_filter()
    wb2 = load_workbook(file)
    print "loading filter ok."
    sheet = sheet_Name
    for i in range(1, wb2[sheet].max_row + 1):
        cellD = "D" + str(i)
        # first7a = re.search(r'\d{7}',wb2["all"]["D1"].value)
        cell_Num = wb2[sheet][cellD].value
        try:
            if re.search(r'\d{11}', cell_Num) and type(int(cell_Num)) == long:
                first7 = re.search(r'\d{7}', str(cell_Num))
                if str(first7.group()) in Tel_Ningbo:
                    print "%8s is a Ningbo Number." %(cell_Num)
                else:
                    pass
               #     print "%8s is not a Ningbo Number." %i
            else:
                pass
               # print "%8s is not a telephone number." %i
        except:
            pass
#            print "%8s have illegal unicode number!" %i
filter_Number("hirun.xlsx","Sheet1")
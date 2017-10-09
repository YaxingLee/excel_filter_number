import os
import threading
from openpyxl import load_workbook
from openpyxl import Workbook
os.chdir(r'd:/xx1')
def convert_N_T(file):
    wb_Con = load_workbook(file)
    for i in range (2,wb_Con["Sheet1"].max_row+1):
        if type(wb_Con["Sheet1"]["B"+str(i)].value) == long:
            pass
        else:
            wb_Con["Sheet1"]["A"+str(i)].value,wb_Con["Sheet1"]["B"+str(i)].value = \
                wb_Con["Sheet1"]["B" + str(i)].value, wb_Con["Sheet1"]["A" + str(i)].value
            print "Exchanging A%s..." %i
    wb_Con.save(file)
def parse_Excel(source,target):
    source_File=source
    target_File=target
    wbs = load_workbook(source)
    wbt = load_workbook(target)
    sheet = wbt.active
    max_Row = wbs["Sheet1"].max_row
    max_Colu = wbs["Sheet1"].max_column
    for i in range(2,max_Row+1):
        sheet.append([wbs["Sheet1"]["A"+str(i)].value,wbs["Sheet1"]["B"+str(i)].value])
    wbt.save(target)
    wbt.close()
    wbs.close()
    print "Appended %s." %source
name_List = os.listdir(".")
threads = []
for i in range(0,len(name_List)):
    threadObj = threading.Thread(target=parse_Excel,args=[name_List[i],"3.xlsx"])
    threads.append(threadObj)
    threadObj.start()

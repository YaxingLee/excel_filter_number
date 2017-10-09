# coding=UTF-8
import os
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
def copy_U_tel(file,sheet):
#return two lists of username and telephone numbers.
    wb1 = load_workbook(file)
    print "loaded %s [+]" %file
    max_ROW = 0
#    print max_ROW
    user_Telnum = []
    user_Name = []
#    for i in wb1[sheet].iter_cols(min_row=2,min_col=2,max_row=max_ROW):
#        for num in i:
#            print type(num.value)
#            if num.value == None:
#                break
#            else:
#                user_Telnum.append(num.value)
#    for t_name in wb1[sheet].iter_cols(min_row=2,min_col=1,max_row=max_ROW):
#        for name in t_name:
#            if name.value == None:
#                break
#            else:
#                user_Name.append(name.value)
    name_A_tel_tuple = tuple(wb1[sheet])
    for i in range(1,len(name_A_tel_tuple)):
        if name_A_tel_tuple[i][0].value != None:
            max_ROW+=1
            user_Telnum.append(name_A_tel_tuple[i][1].value)
            user_Name.append(name_A_tel_tuple[i][0].value)
        else:
            break
    print max_ROW
    if len(user_Name) == len(user_Telnum):
        print "[+] Geted username and telephone number,and available."
    else:
        print "[-] Error at getting username and telephone number."
    return user_Name,user_Telnum

def get_Province_City(user_Telnum):
    url = "http://www.sjgsd.com/n/?q=%s" %user_Telnum
#    print url
    res = requests.get(url).text
    soup = BeautifulSoup(res,"html.parser")
#    print soup.dl
    dl_Lines = soup.dl.contents[3]
#    print dl_Lines.span.next_sibling.next_sibling
    Province = dl_Lines.find_all("a")[1].string
    City = dl_Lines.find_all("a")[2].string
    return Province,City

def create_append_Num(user_Name,user_Telnum,province_Name,city):
    try:
        wb = load_workbook(filename="selected.xlsx")
    except:
        wb = Workbook()
#    print wb.sheetnames
    if province_Name not in wb.sheetnames:
        wb.create_sheet(province_Name)
        wb[province_Name].append([user_Name, user_Telnum, city])
    else:
        wb[province_Name].append([user_Name,user_Telnum,city])
    wb.save("selected.xlsx")
    wb.close()

os.chdir(r'D:\ll\getpro')
name,num = copy_U_tel('hirun.xlsx',"Sheet1")
for i in range(0,len(name)):
    pro,city = get_Province_City(num[i])
    print name[i],num[i],pro,city
    create_append_Num(name[i],num[i],pro,city)
print "done"